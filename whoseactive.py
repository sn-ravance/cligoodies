import openpyxl
import win32com.client

# Open Excel workbook
workbook = openpyxl.load_workbook('path/to/your/file.xlsx')
worksheet = workbook.active

# Create Outlook Application object
outlook = win32com.client.Dispatch("Outlook.Application")

# Iterate through rows until a blank cell is encountered in column B
for row in worksheet.iter_rows(min_row=2, max_col=2):
    email_address = row[0].value

    # Break the loop if email address is blank
    if not email_address:
        break

    # Create Outlook MailItem object
    mail_item = outlook.CreateItem(0)

    # Set email address
    mail_item.To = email_address

    # Try sending the email to check if it's active
    try:
        mail_item.Send()

        # Email address is active
        print(f"Email address {email_address} is active")

    except Exception as e:
        # Email address is disabled
        print(f"Email address {email_address} is disabled")

        # Highlight the row yellow
        for cell in row:
            cell.fill = openpyxl.styles.PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

# Save the modified workbook
workbook.save('path/to/your/updated_file.xlsx')

# Close the workbook
workbook.close()
